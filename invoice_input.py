import pyautogui, openpyxl, datetime, calendar, os, sys, re
import subprocess, time
import pygetwindow as gw
import pywinauto
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date

from tkinter import Tk, simpledialog, messagebox
from tkinter.filedialog import askopenfilename, Frame, Button
from time import sleep
from pyautogui import press, write, hotkey
from quickbooks_helper import find_message_window, handle_error

def quick_windows():
    windows = gw.getAllWindows()

    qb_window = None

    for window in windows:
        if "QuickBooks Desktop" in window.title:
            qb_window = window

    return qb_window       
 
if __name__ == "__main__":
    app = pywinauto.Application(backend="uia").connect(path=r"C:\Program Files (x86)\Intuit\QuickBooks 2019\QBW32.EXE")
    main_window = app.window(title_re='.*QuickBooks Desktop Pro 2019.*')

    file_name = f'{date.today()}_invoices.xlsx'

    folder_path = r'C:\Users\Michael\Desktop\python-work\Invoices'

    invoices_for_input = load_workbook(os.path.join(folder_path, file_name))


    ws = invoices_for_input['Sheet']

    first_row_skipped = False

    # Alright, this starts with the bill window open in QB

    qb_window = quick_windows()
    sleep(1)

    if qb_window:
        qb_window.activate()
    else:
        print("QuickBooks window not found.")   

    sleep(3)

    for i in range(2,len(ws['A'])+1):
        print(ws[f'A{i}'].value)
        print(ws[f'B{i}'].value)
        print(ws[f'C{i}'].value)
        print(ws[f'D{i}'].value)
        sleep(1.5)
        pyautogui.write(ws[f'A{i}'].value)
        sleep(1.5)
        pyautogui.press('tab')
        sleep(1.5)
        pyautogui.write(ws[f'B{i}'].value)
        sleep(1.5)
        pyautogui.press('tab')
        sleep(1.5)
        if ws[f'C{i}'].value is not None:
            pyautogui.write(ws[f'C{i}'].value)
            sleep(1)
        pyautogui.press('tab')
        sleep(1)
        pyautogui.write(str(ws[f'D{i}'].value))
        sleep(1)
        hotkey('alt', 's')
        # sleep(3)
        message_window_exists, message_window = find_message_window(app.top_window())
        if message_window_exists:
            print("Error message detected. Handling...")
            handle_error(message_window, main_window)
        sleep(.5)

    print('Entered!')

